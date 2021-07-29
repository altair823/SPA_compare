from openpyxl import Workbook, load_workbook

filename = "100 200x200_dik^dikpq^dikbq^aspq^asbq_sd2"

write_wb = load_workbook('excel/'+filename+'.xlsx')

write_ws = write_wb.active
write_ws.append(['DIK', 'DIKPQ', 'DIKBQ', 'ASPQ', 'ASBQ', 'DIK', 'DIKPQ', 'DIKBQ', 'ASPQ', 'ASBQ'])
with open('rawtext/' + filename + '.txt') as resultFile:
    lines = resultFile.readlines()
    resultDataList = {}
    resultRound = 0
    resultLength = 0
    resultTime = 0
    resultType = ''
    currentRow = 2
    for line in lines:
        if line[:8] == "[Round: ":
            resultRound = int(line[8:-2])
            continue
        elif line[6:21] == '\'s SP length : ':
            resultType = line[0:5]
            resultType = resultType.strip()
            end = line.find('-')
            resultLength = int(line[21:end-1])
            start = line.find('Execution time is : ')
            start += len('Execution time is : ')
            resultTime = float(line[start:-1])
            resultDataList[resultType] = [resultLength, resultTime]
        elif line[0:1] == '\n':
            if len(resultDataList) == 0:
                continue
            allData = []
            for Type in resultDataList.keys():
                allData.append(resultDataList[Type][0])
            for Type in resultDataList.keys():
                allData.append(resultDataList[Type][1])
            for i in range(0, len(allData)):
                write_ws.cell(row=currentRow, column=i+1).value = allData[i]
            resultDataList.clear()
            currentRow += 1

write_wb.save('excel/' + filename + '.xlsx')
